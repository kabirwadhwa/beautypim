from pathlib import Path

from openpyxl import Workbook

from app.knowledge_corpus.adapters import RetailFeedAdapter, RichBeautyWorkbookAdapter
from app.knowledge_corpus.import_service import create_import_job, import_corpus
from app.knowledge_corpus.normalization import normalized_brand, normalized_gtin, normalized_text, split_size
from app.knowledge_corpus.retrieval import evidence_is_sufficient, resolve_exact_field_evidence, retrieve_corpus_evidence
from app.knowledge_corpus.evaluation import evaluate_holdout, evaluate_non_ean_matching
from app.models import KnowledgeFormulation, KnowledgeMarketObservation, KnowledgeProduct, KnowledgeSourceObservation, KnowledgeVariant


def _token(client, email):
    return client.post("/api/auth/token", data={"username": email, "password": "securepassword123"}).json()["access_token"]


def _retail_fixture(path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(["aw_deep_link", "product_name", "merchant_product_id", "merchant_image_url", "description",
               "merchant_category", "search_price", "merchant_name", "currency", "merchant_deep_link",
               "language", "last_updated", "brand_name", "colour", "product_type",
               "merchant_product_category_path", "rrp_price", "ean", "parent_product_id", "in_stock", "ShoppingNL:size",
               "average_rating", "review_count"])
    ws.append(["https://example.test/p/1", "Evidence Serum 30 ml", "SKU-1", "https://img.test/1.jpg",
               "Hydrating niacinamide serum", "Gezichtsverzorging", "24.95", "Retail", "EUR",
               "https://example.test/p/1", "nl", "2024-11-25", "Proof Lab", "", "Serum",
               "Beauty > Gezichtsverzorging > Serum", "29.95", "8712345678901", "PARENT-1", True, "30 ml", "4.7", 128])
    ws.append(["https://example.test/p/2", "Evidence Serum Rose 30 ml", "SKU-2", "https://img.test/2.jpg",
               "Rose variant", "Gezichtsverzorging", "25.95", "Retail", "EUR", "https://example.test/p/2",
               "nl", "2024-11-25", "Proof Lab", "Rose", "Serum", "Beauty > Gezichtsverzorging > Serum",
               "30.95", "8712345678918", "PARENT-1", False, "30 ml", "4.4", 64])
    wb.save(path)


def _rich_fixture(path: Path):
    wb = Workbook(); descriptions = wb.active; descriptions.title = "PRODUCT OMSCHRIJVINGEN"
    descriptions.append(["SKU number", "Base Product", "Range Name", "Content size", "Content unit",
        "Approvalstatus description", "Product Name nl_BE", "Product Name fr_BE", "Prod.Info.Descript_fr_BE",
        "Prod.Info.Descript_nl_BE", "Approvalstatus description", "Ingredients fr_BE", "Ingredients nl_BE",
        "Uses fr_BE", "Uses nl_BE", "Informativetext1 fr_BE", "Informativetext1 nl_BE",
        "Informativetext2 fr_BE", "Informativetext2 nl_BE", "Informativetext3 fr_BE", "Informativetext3 nl_BE"])
    descriptions.append(["100", "BASE-1", "Hydra", 50, "ml", "Approved", "Hydra Crème", "Crème Hydra", "", "Hydraterende crème", "Approved",
                         "Aqua, Glycerin", "Aqua, Glycerin, Niacinamide", "", "Dagelijks aanbrengen", "", "", "", "", "", ""])
    attributes = wb.create_sheet("PRODUCT ATTRIBUTEN")
    headers = ["Product type", "SKU number", "Base Product", "MAIN BRAND", "Range Name", "Approvalstatus description",
               "Product Classification nl_BE", "Fragrance Family nl_BE", "Fragrance Style nl_BE", "SPF nl_BE", "Benefit nl_BE",
               "Skin Condition Treatment nl_BE", "Skin Type nl_BE", "Age Range nl_BE", "Lifestyle nl_BE", "For Whom nl_BE",
               "Key Ingredient nl_BE", "Durability nl_BE", "Intensity nl_BE", "Content Measure nl_BE", "Color Packaging nl_BE",
               "Shape of the Bottle nl_BE", "Color Tone nl_BE", "Routine nl_BE", "Coverage nl_BE", "Free From nl_BE", "Body Area nl_BE",
               "Hair Type nl_BE", "Hair Color nl_BE", "Hair Use nl_NL", "Product Feature nl_BE", "Moment of Use nl_BE", "Type Dispenser nl_BE",
               "Concern Condition nl_BE", "Undertone nl_BE", "Finish nl_BE", "Product Format nl_BE", "SOORTGROEPNUMMER ",
               "MDC PARFUM", "MDC MAKE UP", "MDC SKINCARE", "MDC BODY", "MDC HAIR", "MDC HOME"]
    attributes.append(headers)
    row = {"Product type": "Crème", "SKU number": "100", "Base Product": "BASE-1", "MAIN BRAND": "Proof Lab",
           "Range Name": "Hydra", "Product Classification nl_BE": "Gezichtsverzorging", "Skin Type nl_BE": "Droog; Gevoelig",
           "Key Ingredient nl_BE": "Niacinamide", "Finish nl_BE": "Dewy", "Product Format nl_BE": "Crème"}
    attributes.append([row.get(header) for header in headers]); wb.save(path)


def test_multilingual_identifier_and_size_normalization():
    assert normalized_gtin("8712345678901.0") == "8712345678901"
    assert normalized_text("Crème Hydratante") == "creme hydratante"
    assert normalized_text("L'Oréal") == normalized_text("L’Oréal") == "loreal"
    assert normalized_brand("YSL") == normalized_brand("Yves Saint Laurent")
    assert normalized_brand("Kiehl's Since 1851") == normalized_brand("Kiehl’s")
    assert split_size("100 ml") == ("100", "ml")
    adapter = RetailFeedAdapter("fixture")
    assert adapter._is_beauty({"merchant_product_category_path": "Makeup > Foundation"})
    assert adapter._is_beauty({"merchant_product_category_path": "Gezondheid > Dieet & levensstijl > Massage"})
    assert adapter._is_beauty({"merchant_product_category_path": "Home & Lifestyle > Tassen & bagage > Toilettassen"})
    assert not adapter._is_beauty({"merchant_product_category_path": "Home & Lifestyle > Sieraden > Ringen"})


def test_missing_stock_is_not_falsely_recorded_as_out_of_stock(tmp_path):
    path = tmp_path / "retail-missing-stock.xlsx"; _retail_fixture(path)
    from openpyxl import load_workbook
    workbook = load_workbook(path); sheet = workbook.active
    stock_column = [cell.value for cell in sheet[1]].index("in_stock") + 1
    sheet.cell(row=2, column=stock_column).value = None; workbook.save(path)
    record = next(RetailFeedAdapter("fixture").iter_records(str(path)))
    assert record.availability is None


def test_retail_parent_variants_and_exact_retrieval(db, tmp_path):
    path = tmp_path / "retail.xlsx"; _retail_fixture(path)
    adapter = RetailFeedAdapter("fixture_feed")
    job = create_import_job(db, str(path), adapter, "fixture_feed", "Retail Data")
    import_corpus(db, job, str(path), adapter)
    assert db.query(KnowledgeSourceObservation).count() == 2
    assert db.query(KnowledgeMarketObservation).count() == 2
    result = retrieve_corpus_evidence(db, gtin="8712345678901", brand="wrong", product_name="wrong")
    assert result["match_level"] == "exact_product"
    assert result["exact_matches"][0]["gtin"] == "8712345678901"
    assert evidence_is_sufficient(result)
    exact = resolve_exact_field_evidence(result)
    assert float(exact["values"]["rating"]) == 4.7
    assert exact["values"]["review_count"] == 128
    assert exact["market"]["image_url"] == "https://img.test/1.jpg"


def test_weak_same_category_row_is_not_returned_as_comparable(db, tmp_path):
    path = tmp_path / "retail.xlsx"; _retail_fixture(path)
    adapter = RetailFeedAdapter("fixture_feed")
    import_corpus(db, create_import_job(db, str(path), adapter, "fixture_feed", "Retail Data"), str(path), adapter)
    result = retrieve_corpus_evidence(
        db, brand="Different Brand", product_name="Unrelated Night Balm", category="Skin Care",
    )
    assert result["match_level"] == "unmatched"
    assert result["diagnostics"]["qualified_candidate_count"] == 0


def test_exact_ean_bridge_reconciles_cross_dataset_parent_families(db, tmp_path):
    first = tmp_path / "first.xlsx"; second = tmp_path / "second.xlsx"
    _retail_fixture(first); _retail_fixture(second)
    from openpyxl import load_workbook
    workbook = load_workbook(second); sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    sheet.cell(row=2, column=headers.index("parent_product_id") + 1).value = "SECOND-PARENT"
    sheet.cell(row=3, column=headers.index("parent_product_id") + 1).value = "SECOND-PARENT"
    sheet.cell(row=3, column=headers.index("ean") + 1).value = "8712345678994"
    sheet.cell(row=3, column=headers.index("merchant_product_id") + 1).value = "SECOND-SKU"
    workbook.save(second)
    for path, dataset in ((first, "feed_a"), (second, "feed_b")):
        adapter = RetailFeedAdapter(dataset)
        import_corpus(db, create_import_job(db, str(path), adapter, dataset, "Retail Data"), str(path), adapter)
    assert db.query(KnowledgeProduct).count() == 1
    assert db.query(KnowledgeVariant).count() == 3
    assert db.query(KnowledgeSourceObservation).count() == 4


def test_rich_workbook_join_formulation_and_idempotency(db, tmp_path):
    path = tmp_path / "rich.xlsx"; _rich_fixture(path)
    adapter = RichBeautyWorkbookAdapter()
    job = create_import_job(db, str(path), adapter, "rich_beauty_reference", "Retail Data")
    import_corpus(db, job, str(path), adapter)
    assert db.query(KnowledgeFormulation).count() == 1
    assert db.query(KnowledgeFormulation).first().normalized_ingredients[2]["normalized_name"] == "niacinamide"
    same_job = create_import_job(db, str(path), adapter, "rich_beauty_reference", "Retail Data")
    assert same_job.id == job.id
    import_corpus(db, same_job, str(path), adapter)
    assert db.query(KnowledgeSourceObservation).count() == 1


def test_rich_workbook_skus_never_share_formulation_on_weak_range_identity(db, tmp_path):
    path = tmp_path / "rich.xlsx"; _rich_fixture(path)
    from openpyxl import load_workbook
    workbook = load_workbook(path)
    descriptions = workbook["PRODUCT OMSCHRIJVINGEN"]
    descriptions.append(["101", "BASE-1", "Hydra", 50, "ml", "Approved", "Hydra Crème", "Crème Hydra", "", "Hydraterende crème", "Approved",
                         "Aqua, Retinol", "Aqua, Retinol", "", "Dagelijks aanbrengen", "", "", "", "", "", ""])
    attributes = workbook["PRODUCT ATTRIBUTEN"]
    values = [cell.value for cell in attributes[2]]
    sku_column = [cell.value for cell in attributes[1]].index("SKU number")
    values[sku_column] = "101"; attributes.append(values); workbook.save(path)
    adapter = RichBeautyWorkbookAdapter()
    import_corpus(db, create_import_job(db, str(path), adapter, "rich_beauty_reference", "Retail Data"), str(path), adapter)
    assert db.query(KnowledgeVariant).count() == 2
    assert db.query(KnowledgeFormulation).count() == 2
    assert len({row.knowledge_variant_id for row in db.query(KnowledgeFormulation).all()}) == 2


def test_comparable_retrieval_never_exposes_direct_only_evidence(db, tmp_path):
    path = tmp_path / "retail.xlsx"; _retail_fixture(path)
    adapter = RetailFeedAdapter("fixture_feed")
    import_corpus(db, create_import_job(db, str(path), adapter, "fixture_feed", "Retail Data"), str(path), adapter)
    result = retrieve_corpus_evidence(db, product_name="Unrelated Hydrating Toner", category="Skin Care")
    for item in result.get("comparables", []):
        assert not ({"gtin", "claims", "raw_inci", "price", "availability"} & set(item["fields"]))
        assert not item["formulations"]
        assert not item["market_observations"]


def test_conflicted_exact_identity_never_skips_research(db, tmp_path):
    path = tmp_path / "retail.xlsx"; _retail_fixture(path)
    from openpyxl import load_workbook
    workbook = load_workbook(path); sheet = workbook.active
    ean_column = [cell.value for cell in sheet[1]].index("ean") + 1
    sheet.cell(row=3, column=ean_column).value = "8712345678901"
    sheet.cell(row=3, column=2).value = "Different Bundle Identity"
    workbook.save(path)
    adapter = RetailFeedAdapter("fixture_feed")
    import_corpus(db, create_import_job(db, str(path), adapter, "fixture_feed", "Retail Data"), str(path), adapter)
    result = retrieve_corpus_evidence(db, gtin="8712345678901")
    assert result["match_level"] == "conflict"
    assert "product_name" in {row["field_name"] for row in result["exact_matches"][0]["conflicts"]}
    assert not evidence_is_sufficient(result)


def test_holdout_evaluation_measures_exact_identity(db, tmp_path):
    path = tmp_path / "retail.xlsx"; _retail_fixture(path)
    adapter = RetailFeedAdapter("fixture_feed")
    import_corpus(db, create_import_job(db, str(path), adapter, "fixture_feed", "Retail Data"), str(path), adapter)
    report = evaluate_holdout(db, sample_size=2)
    assert report["exact_gtin_lookup_success_rate"] == 1
    assert report["correct_variant_match_rate"] == 1
    assert report["incorrect_match_rate"] == 0


def test_non_ean_evaluation_and_ambiguous_name_safety(db, tmp_path):
    path = tmp_path / "rich.xlsx"; _rich_fixture(path)
    adapter = RichBeautyWorkbookAdapter()
    import_corpus(db, create_import_job(db, str(path), adapter, "rich_beauty_reference", "Retail Data"), str(path), adapter)
    report = evaluate_non_ean_matching(db, sample_size=1)
    assert report["exact_source_id"]["correct_rate"] == 1
    result = retrieve_corpus_evidence(db, brand="Proof Lab", product_name="Hydra Crème", size="50 ml")
    assert result["match_level"] == "exact_product"


def test_corpus_api_permissions(client):
    viewer = {"Authorization": f"Bearer {_token(client, 'viewer@test.com')}"}
    admin = {"Authorization": f"Bearer {_token(client, 'admin@test.com')}"}
    assert client.get("/api/knowledge-corpus/metrics", headers=viewer).status_code == 200
    assert client.get("/api/knowledge-corpus/imports", headers=viewer).status_code == 403
    assert client.get("/api/knowledge-corpus/imports", headers=admin).status_code == 200
    assert client.get("/api/knowledge-corpus/search", headers=viewer).status_code == 403
    assert client.get("/api/knowledge-corpus/search", headers=admin).status_code == 422

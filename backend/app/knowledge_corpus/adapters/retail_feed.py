from __future__ import annotations

from datetime import datetime
from typing import Any, Iterable

from openpyxl import load_workbook

from app.knowledge_corpus.normalization import clean_text, decimal_value, normalized_category, normalized_gtin, normalized_identifier, split_size
from app.knowledge_corpus.records import CorpusAdapter, CorpusRecord


class RetailFeedAdapter(CorpusAdapter):
    name = "retail_feed"
    version = "1.1"

    def __init__(self, dataset_key: str, *, market: str = "NL"):
        self.dataset_key = dataset_key
        self.market = market

    @staticmethod
    def _is_beauty(row: dict[str, Any]) -> bool:
        taxonomy = " ".join(clean_text(row.get(key)).lower() for key in (
            "merchant_category", "category_name", "merchant_product_category_path",
            "merchant_product_second_category", "merchant_product_third_category", "product_type",
        ))
        beauty_terms = (
            "beauty", "cosmetic", "makeup", "make-up", "maquillage", "parfum", "perfume", "fragrance",
            "skin", "huid", "gezicht", "visage", "hair", "haar", "cheveux", "body care", "lichaam",
            "corps", "nail", "nagel", "lip", "eye", "oog", "wenkbrauw", "brow", "mascara",
            "foundation", "concealer", "serum", "cream", "creme", "crème", "cleanser", "shampoo",
            "conditioner", "deodorant", "bath", "douche", "sun care", "zonne", "grooming",
            "scheer", "oral care", "mondverzorging", "sets & kits",
            # Beauty-adjacent departments that are valid catalogue knowledge
            # even though they do not fit the four primary enrichment modules.
            "toilettas", "badkameraccessoires", "massage", "gua sha", "scalp brush",
            "geurstokjes", "roomspray", "kaarsen", "home fragrance", "slaapmasker",
        )
        if any(term in taxonomy for term in beauty_terms):
            return True
        # Retail feeds occasionally place genuine beauty products in seasonal
        # or generic lifestyle departments. Only strong product-name signals
        # can override an otherwise non-beauty taxonomy.
        name = clean_text(row.get("product_name")).lower()
        strong_name_terms = (
            "eau de parfum", "eau de toilette", "body spray", "face wash",
            "hair towel", "hair wrap", "scalp brush", "nipple cream",
            "make-up organizer", "makeup organizer", "make-up advent",
            "beauty advent", "body care advent", "brow lift treatment",
            "gua sha", "makeup applicator", "make-up applicator",
            "zuignap spiegel", "cosmeticaspiegel",
        )
        return any(term in name for term in strong_name_terms)

    @staticmethod
    def _rows(sheet):
        rows = sheet.iter_rows(values_only=True)
        headers = [clean_text(value) for value in next(rows)]
        # Some exported sheets contain a corrupted A1 header. Its values are URLs.
        if headers and headers[0] != "aw_deep_link":
            headers[0] = "aw_deep_link"
        for number, values in enumerate(rows, start=2):
            yield number, {headers[index]: values[index] for index in range(min(len(headers), len(values)))}

    def inspect(self, path: str) -> dict[str, Any]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        inventory = {}
        for sheet in workbook.worksheets:
            rows = sheet.max_row
            columns = sheet.max_column
            if rows is None or columns is None:
                iterator = sheet.iter_rows(values_only=True)
                header = next(iterator, ())
                rows = 1 + sum(1 for _ in iterator)
                columns = len(header)
            inventory[sheet.title] = {"rows": max(0, rows - 1), "columns": columns}
        return inventory

    def iter_records(self, path: str, *, start_row: int = 0, limit: int | None = None) -> Iterable[CorpusRecord]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        yielded = 0
        for offset, (number, row) in enumerate(self._rows(sheet)):
            if offset < start_row:
                continue
            if not self._is_beauty(row):
                yield CorpusRecord(
                    dataset_key=self.dataset_key, sheet=sheet.title, row_number=number,
                    raw_payload=row, source_record_id=normalized_identifier(row.get("merchant_product_id") or row.get("aw_product_id")),
                    source_parent_id=normalized_identifier(row.get("parent_product_id")),
                    brand=clean_text(row.get("brand_name")) or "Unknown Brand",
                    product_name=clean_text(row.get("product_name")) or "Excluded reference row",
                    skip_reason="outside_beauty_domain",
                )
                yielded += 1
                if limit is not None and yielded >= limit:
                    return
                continue
            brand = clean_text(row.get("brand_name")) or "Unknown Brand"
            name = clean_text(row.get("product_name"))
            record_id = normalized_identifier(row.get("merchant_product_id") or row.get("aw_product_id"))
            if not name or not record_id:
                continue
            size_value, size_unit = split_size(row.get("ShoppingNL:size") or row.get("Fashion:size"))
            category_path = clean_text(row.get("merchant_product_category_path") or row.get("category_name") or row.get("merchant_category"))
            last_updated = row.get("last_updated")
            if isinstance(last_updated, str):
                try: last_updated = datetime.fromisoformat(last_updated.replace("Z", "+00:00"))
                except ValueError: last_updated = None
            fields = {
                "description": clean_text(row.get("description")),
                "category_path": category_path,
                "colour": clean_text(row.get("colour")),
                "product_type": clean_text(row.get("product_type")),
                "specifications": clean_text(row.get("specifications")),
                "images": [clean_text(row.get(key)) for key in ("merchant_image_url", "large_image", "alternate_image", "alternate_image_two", "alternate_image_three", "alternate_image_four") if clean_text(row.get(key))],
                "rating": decimal_value(row.get("rating") or row.get("average_rating") or row.get("review_rating")),
                "review_count": row.get("review_count") or row.get("reviews_count") or row.get("number_of_reviews"),
                "review_summary": clean_text(row.get("review_summary") or row.get("reviews_summary")),
            }
            fields = {key: value for key, value in fields.items() if value not in (None, "", [], {})}
            stock_value = row.get("in_stock")
            normalized_stock = str(stock_value).strip().lower() if stock_value is not None else ""
            availability = None
            if normalized_stock in {"1", "true", "yes", "ja", "in stock", "in_stock"}:
                availability = "in_stock"
            elif normalized_stock in {"0", "false", "no", "nee", "out of stock", "out_of_stock"}:
                availability = "out_of_stock"
            yield CorpusRecord(
                dataset_key=self.dataset_key, sheet=sheet.title, row_number=number, raw_payload=row,
                source_record_id=record_id, source_parent_id=normalized_identifier(row.get("parent_product_id")),
                brand=brand, product_name=name, variant_name=clean_text(row.get("colour") or row.get("ShoppingNL:size") or row.get("Fashion:size")) or None,
                gtin=normalized_gtin(row.get("ean") or row.get("product_GTIN") or row.get("upc")),
                size_value=size_value, size_unit=size_unit, shade=clean_text(row.get("colour")) or None,
                colour=clean_text(row.get("colour")) or None, category=normalized_category(category_path),
                subcategory=category_path or None, product_type=clean_text(row.get("product_type")) or None,
                locale=clean_text(row.get("language")) or "nl", market=self.market,
                source_retailer="Retail Data", source_url=clean_text(row.get("merchant_deep_link") or row.get("aw_deep_link")) or None,
                observed_at=last_updated if isinstance(last_updated, datetime) else None, fields=fields,
                price=decimal_value(row.get("search_price") or row.get("store_price")),
                original_price=decimal_value(row.get("rrp_price")), currency=clean_text(row.get("currency")) or None,
                availability=availability,
                image_url=clean_text(row.get("merchant_image_url") or row.get("large_image")) or None,
            )
            yielded += 1
            if limit is not None and yielded >= limit:
                return

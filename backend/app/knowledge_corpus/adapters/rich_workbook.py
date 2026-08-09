from __future__ import annotations

from collections import defaultdict
from typing import Any, Iterable

from openpyxl import load_workbook

from app.knowledge_corpus.normalization import clean_text, list_value, normalized_category, normalized_identifier, split_size
from app.knowledge_corpus.records import CorpusAdapter, CorpusRecord


class RichBeautyWorkbookAdapter(CorpusAdapter):
    name = "rich_beauty_workbook"
    version = "1.1"
    descriptions_sheet = "PRODUCT OMSCHRIJVINGEN"
    attributes_sheet = "PRODUCT ATTRIBUTEN"

    @staticmethod
    def _rows(sheet):
        rows = sheet.iter_rows(values_only=True)
        headers = [clean_text(value) for value in next(rows)]
        for number, values in enumerate(rows, start=2):
            yield number, {headers[index]: values[index] for index in range(min(len(headers), len(values)))}

    def inspect(self, path: str) -> dict[str, Any]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        return {sheet.title: {"rows": sheet.max_row - 1, "columns": sheet.max_column} for sheet in workbook.worksheets}

    def iter_records(self, path: str, *, start_row: int = 0, limit: int | None = None) -> Iterable[CorpusRecord]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        descriptions: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
        for number, row in self._rows(workbook[self.descriptions_sheet]):
            sku = normalized_identifier(row.get("SKU number"))
            if sku:
                descriptions[sku].append((number, row))

        yielded = 0
        for offset, (number, attributes) in enumerate(self._rows(workbook[self.attributes_sheet])):
            if offset < start_row:
                continue
            sku = normalized_identifier(attributes.get("SKU number"))
            candidates = descriptions.get(sku or "") or [(number, {})]
            desc_number, description = candidates[0]
            brand = clean_text(attributes.get("MAIN BRAND"))
            product_name = clean_text(description.get("Product Name nl_BE") or description.get("Product Name fr_BE") or attributes.get("Range Name"))
            if not brand or not product_name:
                yield CorpusRecord(
                    dataset_key="rich_beauty_reference", sheet=self.attributes_sheet, row_number=number,
                    raw_payload={"attributes": attributes, "description": description, "description_source_row": desc_number},
                    source_record_id=sku, source_parent_id=normalized_identifier(attributes.get("Base Product")),
                    brand=brand or "Unknown Brand", product_name=product_name or "Unidentified reference row",
                    skip_reason="missing_usable_identity",
                )
                yielded += 1
                if limit is not None and yielded >= limit:
                    return
                continue
            size_value, size_unit = split_size(description.get("Content size"), description.get("Content unit"))
            classification = clean_text(attributes.get("Product Classification nl_BE"))
            fields = {
                "description": clean_text(description.get("Prod.Info.Descript_nl_BE") or description.get("Prod.Info.Descript_fr_BE")),
                "directions": clean_text(description.get("Uses nl_BE") or description.get("Uses fr_BE")),
                "benefits": list_value(attributes.get("Benefit nl_BE")),
                "targeted_concerns": list_value(attributes.get("Skin Condition Treatment nl_BE") or attributes.get("Concern Condition nl_BE")),
                "skin_types": list_value(attributes.get("Skin Type nl_BE")),
                "hair_types": list_value(attributes.get("Hair Type nl_BE")),
                "key_ingredients": list_value(attributes.get("Key Ingredient nl_BE")),
                "claims": list_value(attributes.get("Free From nl_BE")) + list_value(attributes.get("Product Feature nl_BE")),
                "finish": clean_text(attributes.get("Finish nl_BE")),
                "texture_format": clean_text(attributes.get("Product Format nl_BE")),
                "coverage": clean_text(attributes.get("Coverage nl_BE")),
                "fragrance_family": clean_text(attributes.get("Fragrance Family nl_BE")),
                "fragrance_style": clean_text(attributes.get("Fragrance Style nl_BE")),
                "longevity": clean_text(attributes.get("Durability nl_BE")),
                "sillage_projection": clean_text(attributes.get("Intensity nl_BE")),
                "application_area": clean_text(attributes.get("Body Area nl_BE")),
                "routine": clean_text(attributes.get("Routine nl_BE") or attributes.get("Moment of Use nl_BE")),
                "informative_text": [clean_text(description.get(f"Informativetext{i} nl_BE") or description.get(f"Informativetext{i} fr_BE")) for i in range(1, 4)],
            }
            fields = {key: value for key, value in fields.items() if value not in (None, "", [], {})}
            raw = {"attributes": attributes, "description": description, "description_source_row": desc_number}
            yield CorpusRecord(
                dataset_key="rich_beauty_reference", sheet=self.attributes_sheet, row_number=number,
                raw_payload=raw, source_record_id=sku, source_parent_id=normalized_identifier(attributes.get("Base Product")),
                brand=brand, product_name=product_name, variant_name=clean_text(attributes.get("Range Name")) or None,
                size_value=size_value, size_unit=size_unit, category=normalized_category(classification),
                subcategory=classification or None, product_type=clean_text(attributes.get("Product type")) or None,
                application_area=fields.get("application_area"), locale="nl-BE/fr-BE", market="BE",
                source_retailer="Retail Data", fields=fields,
                raw_inci=clean_text(description.get("Ingredients nl_BE") or description.get("Ingredients fr_BE")) or None,
                inci_language="nl-BE" if description.get("Ingredients nl_BE") else "fr-BE",
            )
            yielded += 1
            if limit is not None and yielded >= limit:
                return
